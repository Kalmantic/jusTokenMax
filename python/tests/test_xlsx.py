"""XLSX -> Markdown digest (optimize kind="xlsx").

Builds real .xlsx workbooks with openpyxl at runtime (no committed binaries,
matching the conftest convention) and asserts the schema+sample contract. This
is the one Office format that genuinely compresses: a big sheet collapses to a
schema + a handful of sample rows, exactly like the CSV handler.
"""

import json

import pytest

# openpyxl is the optional `office` extra; skip (don't error) when absent.
pytest.importorskip("openpyxl")

from justokenmax.xlsx import xlsx_to_markdown
from justokenmax.optimize import optimize


def _wb(path, rows=100, sheets=("Data",), cols=("id", "name", "active")):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for s in sheets:
        ws = wb.create_sheet(s)
        ws.append(list(cols))
        for i in range(rows):
            ws.append([i, f"user{i}", i % 2 == 0])
    wb.save(str(path))


def test_schema_and_counts(tmp_path):
    p = tmp_path / "book.xlsx"
    _wb(p, rows=100)
    digest, full, stats = xlsx_to_markdown(str(p))
    assert "## Sheet: Data" in digest
    assert "100 rows" in digest
    assert "id: int" in digest
    assert "name: str" in digest
    assert "active: bool" in digest
    assert stats["sheets"] == 1
    assert stats["total_rows"] == 100


def test_samples_and_shrinks(tmp_path):
    p = tmp_path / "big.xlsx"
    _wb(p, rows=500)
    digest, full, stats = xlsx_to_markdown(str(p))
    # The digest must be much smaller than rendering every row (real compression).
    assert len(digest) < len(full)
    assert "user0" in digest        # head sample
    assert "user499" in digest      # tail sample
    assert "500 rows" in digest


def test_multiple_sheets(tmp_path):
    p = tmp_path / "multi.xlsx"
    _wb(p, rows=10, sheets=("Alpha", "Beta"))
    digest, _, stats = xlsx_to_markdown(str(p))
    assert "## Sheet: Alpha" in digest
    assert "## Sheet: Beta" in digest
    assert stats["sheets"] == 2


def test_formula_uses_data_only(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws.append(["a", "b", "total"])
    ws.append([111, 222, "=A2+B2"])
    p = tmp_path / "formula.xlsx"
    wb.save(str(p))

    digest, _, _ = xlsx_to_markdown(str(p))
    # Positive: literal cell values are surfaced.
    assert "111" in digest and "222" in digest
    # Negative: data_only=True means we never emit the raw formula text. (We
    # can only verify the negative here: an openpyxl-authored file has no cached
    # value, so a formula cell reads as None/blank. Asserting the *computed*
    # value would require a real Excel-calculated file — a committed binary,
    # which this project's no-binary-fixtures convention forbids. data_only is
    # still proven in effect: without it openpyxl would return "=A2+B2".)
    assert "=A2+B2" not in digest
    assert "=A2" not in digest


def test_wide_sheet_caps_columns(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Wide"
    headers = [f"c{i}" for i in range(30)]
    ws.append(headers)
    ws.append(list(range(30)))
    p = tmp_path / "wide.xlsx"
    wb.save(str(p))

    digest, _, _ = xlsx_to_markdown(str(p))
    assert "columns not shown" in digest  # truncation notice
    assert "c29" not in digest            # last column dropped by the cap


def test_flags_images_not_extracted(tmp_path):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image

    img_path = tmp_path / "chart.png"
    Image.new("RGB", (16, 16), (10, 80, 200)).save(str(img_path))
    wb = Workbook()
    ws = wb.active
    ws.title = "WithChart"
    ws.append(["x", "y"])
    ws.append([1, 2])
    ws.add_image(XLImage(str(img_path)), "D1")
    p = tmp_path / "imaged.xlsx"
    wb.save(str(p))

    digest, _, stats = xlsx_to_markdown(str(p))
    assert "image(s) in this workbook not extracted" in digest
    assert stats["images"] >= 1


# ---- optimize() dispatch integration ----

def test_optimize_detects_and_compresses_xlsx(tmp_path):
    p = tmp_path / "big.xlsx"
    _wb(p, rows=2000)
    res = optimize(str(p))
    assert res.ok and res.kind == "xlsx"
    assert res.output.endswith(".xlsx.md")
    # Unlike DOCX, XLSX is real compression: sampled digest << full table.
    assert res.tokens_after < res.tokens_before


def test_secret_in_xlsx_is_masked(tmp_path):
    from openpyxl import Workbook

    secret = "AK" + "IA" + "S" * 16
    wb = Workbook()
    ws = wb.active
    ws.append(["note", "value"])
    ws.append(["aws_key", secret])
    # pad so it clears the min-bytes threshold
    for i in range(50):
        ws.append([f"row{i}", i])
    p = tmp_path / "leaky.xlsx"
    wb.save(str(p))

    res = optimize(str(p))
    assert res.ok
    artifact = open(res.output, encoding="utf-8").read()
    assert secret not in artifact


def test_fail_open_when_library_missing(tmp_path, monkeypatch):
    from justokenmax import xlsx as xlsx_mod

    def _raise_import(_path):
        raise ImportError("No module named 'openpyxl'")

    monkeypatch.setattr(xlsx_mod, "xlsx_to_markdown", _raise_import)
    p = tmp_path / "book.xlsx"
    _wb(p, rows=200)
    res = optimize(str(p))
    assert res.ok is False and res.kind == "skip"
    assert "openpyxl" in res.note


def test_fail_open_on_parse_error(tmp_path, monkeypatch):
    from justokenmax import xlsx as xlsx_mod

    def _raise(_path):
        raise ValueError("bad workbook")

    monkeypatch.setattr(xlsx_mod, "xlsx_to_markdown", _raise)
    p = tmp_path / "book.xlsx"
    _wb(p, rows=200)
    res = optimize(str(p))
    assert res.ok is False and res.kind == "skip"


def test_cli_xlsx_subcommand(tmp_path, capsys):
    from justokenmax.cli import main

    p = tmp_path / "book.xlsx"
    _wb(p, rows=200)
    rc = main(["xlsx", "--json", str(p)])
    out = json.loads(capsys.readouterr().out)
    assert rc == 0
    assert out["ok"] is True and out["kind"] == "xlsx"


# ---- safety caps & edge cases ----

def test_output_size_cap_truncates(tmp_path, monkeypatch):
    from justokenmax import xlsx as xlsx_mod

    monkeypatch.setattr(xlsx_mod, "MAX_OUTPUT_CHARS", 50)
    p = tmp_path / "big.xlsx"
    _wb(p, rows=200)
    digest, _, _ = xlsx_to_markdown(str(p))
    assert "truncated" in digest


def test_empty_and_zero_row_sheets(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Empty")                 # no rows at all
    ws = wb.create_sheet("HeaderOnly")       # header, zero data rows
    ws.append(["id", "name"])
    data = wb.create_sheet("Data")
    data.append(["id", "name"])
    data.append([1, "a"])
    p = tmp_path / "edges.xlsx"
    wb.save(str(p))

    digest, _, stats = xlsx_to_markdown(str(p))  # must not raise
    assert "## Sheet: HeaderOnly" in digest
    assert "0 rows" in digest                    # zero-row sheet rendered honestly
    assert "## Sheet: Data" in digest
    assert stats["total_rows"] == 1              # only Data's one data row counts


def test_image_count_is_exact(tmp_path):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image

    a = tmp_path / "a.png"
    b = tmp_path / "b.png"
    Image.new("RGB", (8, 8), (255, 0, 0)).save(str(a))
    Image.new("RGB", (8, 8), (0, 0, 255)).save(str(b))
    wb = Workbook()
    ws = wb.active
    ws.append(["x"])
    ws.add_image(XLImage(str(a)), "C1")
    ws.add_image(XLImage(str(b)), "C10")
    p = tmp_path / "two_imgs.xlsx"
    wb.save(str(p))

    _, _, stats = xlsx_to_markdown(str(p))
    assert stats["images"] == 2


def test_real_corrupt_file_fails_open(tmp_path):
    # A genuinely broken .xlsx (not a monkeypatched raise): openpyxl raises on
    # load, and optimize() must skip rather than crash the Read hook.
    p = tmp_path / "broken.xlsx"
    p.write_bytes(b"PK\x03\x04 not actually a valid workbook " * 200)
    res = optimize(str(p))
    assert res.ok is False and res.kind == "skip"
